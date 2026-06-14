import csv
import io
from pathlib import Path
from uuid import UUID

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import get_settings
from app.models import Submission, SubmissionFile, UPLOAD_KEYS
from app.services.ingest import clear_ingested_rows, ingest_csv_rows
from app.template_schema import TEMPLATE_HEADERS


def _check_columns(got_headers: list[str], upload_key: str) -> dict:
    """Check that all required columns for upload_key are present (case-insensitive).
    Extra columns are ignored."""
    expected = TEMPLATE_HEADERS.get(upload_key)
    if not expected:
        return {"ok": False, "message": "Unknown template type"}
    got_lower = [h.lower() for h in got_headers]
    missing = [h for h in expected if h.lower() not in got_lower]
    if missing:
        return {
            "ok": False,
            "message": f"Missing required column{'s' if len(missing) > 1 else ''}: {', '.join(missing)}.",
        }
    return {"ok": True, "message": "All required columns found."}


def validate_csv_headers(content: bytes, upload_key: str) -> dict:
    try:
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        row = next(reader, None)
        if not row:
            return {"ok": False, "message": "The file looks empty — no header row found."}
        got = [c.strip().replace("\ufeff", "") for c in row if c.strip()]
        return _check_columns(got, upload_key)
    except Exception:
        return {"ok": False, "message": "Could not parse CSV headers."}


def validate_excel_headers(content: bytes, upload_key: str) -> dict:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return {"ok": False, "message": "The workbook has no sheets."}
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        wb.close()
        if not first_row:
            return {"ok": False, "message": "The file looks empty — no header row found."}
        got = [str(c).strip() for c in first_row if c is not None and str(c).strip()]
        return _check_columns(got, upload_key)
    except Exception:
        return {"ok": False, "message": "Could not read the Excel file. Make sure it is a valid .xlsx."}


async def save_submission_file(
    db: Session,
    submission: Submission,
    upload_key: str,
    file: UploadFile,
) -> SubmissionFile:
    if upload_key not in UPLOAD_KEYS:
        raise HTTPException(status_code=400, detail=f"Invalid upload_key. Use one of: {UPLOAD_KEYS}")

    settings = get_settings()
    upload_root = Path(settings.upload_dir) / str(submission.id) / upload_key
    upload_root.mkdir(parents=True, exist_ok=True)

    content = await file.read()
    safe_name = Path(file.filename or "upload.bin").name
    dest = upload_root / safe_name
    dest.write_bytes(content)

    validation: dict
    name_lower = safe_name.lower()
    if name_lower.endswith(".csv"):
        validation = validate_csv_headers(content, upload_key)
    elif name_lower.endswith(".xlsx"):
        validation = validate_excel_headers(content, upload_key)
    elif name_lower.endswith(".xls"):
        validation = {
            "ok": None,
            "message": "Legacy .xls format — please save as .xlsx or .csv for validation.",
        }
    else:
        validation = {
            "ok": None,
            "message": "Save as .csv or .xlsx to validate headers automatically.",
        }

    existing = db.execute(
        select(SubmissionFile).where(
            SubmissionFile.submission_id == submission.id,
            SubmissionFile.upload_key == upload_key,
        )
    ).scalars().first()
    if existing:
        try:
            Path(existing.storage_path).unlink(missing_ok=True)
        except OSError:
            pass
        clear_ingested_rows(db, submission.id, upload_key)
        db.delete(existing)

    record = SubmissionFile(
        submission_id=submission.id,
        upload_key=upload_key,
        file_name=safe_name,
        content_type=file.content_type,
        size=len(content),
        storage_path=str(dest),
        content=content,
        validation=validation,
    )
    db.add(record)
    db.flush()

    if name_lower.endswith(".csv") and validation.get("ok") is True:
        ingest_csv_rows(db, submission, upload_key, content, record.id)

    meta = submission.uploads_meta or {}
    meta[upload_key] = {
        "fileName": safe_name,
        "size": len(content),
        "savedAt": record.uploaded_at.isoformat() if record.uploaded_at else None,
        "validation": validation,
    }
    submission.uploads_meta = meta
    db.commit()
    db.refresh(record)
    return record


def delete_submission_file(db: Session, submission: Submission, upload_key: str) -> None:
    record = db.execute(
        select(SubmissionFile).where(
            SubmissionFile.submission_id == submission.id,
            SubmissionFile.upload_key == upload_key,
        )
    ).scalars().first()
    if record:
        try:
            Path(record.storage_path).unlink(missing_ok=True)
        except OSError:
            pass
        db.delete(record)
    clear_ingested_rows(db, submission.id, upload_key)
    meta = submission.uploads_meta or {}
    meta[upload_key] = None
    submission.uploads_meta = meta
    db.commit()
