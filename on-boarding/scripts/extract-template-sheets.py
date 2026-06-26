#!/usr/bin/env python3
"""Extract individual data sheets from Helix App Required Templates.xlsx."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / "Helix App Required Templates.xlsx"
OUT_DIR = Path(__file__).resolve().parents[1] / "assets" / "templates"

# template id -> workbook sheet name (matched after strip, case-insensitive)
SHEETS = {
    "departments": "Department",
    "units": "Units",
    "roles": "Roles",
    "staff": "Staff",
    "patients": "Patients",
}


def _find_sheet(workbook, expected: str):
    target = expected.strip().lower()
    for name in workbook.sheetnames:
        if name.strip().lower() == target:
            return workbook[name]
    raise KeyError(f"Sheet not found for {expected!r}: {workbook.sheetnames}")


def _copy_sheet(source_ws, dest_ws) -> None:
    for row in source_ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                dest_ws[cell.coordinate].value = cell.value

    for col, dim in source_ws.column_dimensions.items():
        if dim.width:
            dest_ws.column_dimensions[col].width = dim.width
    for row, dim in source_ws.row_dimensions.items():
        if dim.height:
            dest_ws.row_dimensions[row].height = dim.height


def extract_sheet(workbook, sheet_key: str, expected_name: str) -> Path:
    source_ws = _find_sheet(workbook, expected_name)
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = expected_name.strip()
    _copy_sheet(source_ws, out_ws)
    out_path = OUT_DIR / f"helix-{sheet_key}-template.xlsx"
    out_wb.save(out_path)
    return out_path


def main() -> None:
    if not SOURCE.is_file():
        raise SystemExit(f"Source workbook not found: {SOURCE}")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(SOURCE)
    written = []
    for key, sheet_name in SHEETS.items():
        path = extract_sheet(wb, key, sheet_name)
        written.append(path)
    wb.close()
    for path in written:
        print(path.relative_to(ROOT))


if __name__ == "__main__":
    main()
